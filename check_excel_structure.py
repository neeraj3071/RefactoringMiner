import openpyxl

wb = openpyxl.load_workbook('Final_RQ1&RQ2&RQ3_Updated_Final_Commit_Analysis_With_Tool_Results_Final_Iternation.xlsx')
ws = wb.active

# Check the header row
print("="*80)
print("ALL COLUMN HEADERS:")
print("="*80)
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col_idx).value
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    print(f"Column {col_idx} ({col_letter}): {header}")

print("\n" + "="*80)
print("SAMPLE ROW 2 DATA:")
print("="*80)

for col_idx in range(1, min(15, ws.max_column + 1)):
    header = ws.cell(row=1, column=col_idx).value
    value = ws.cell(row=2, column=col_idx).value
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    if value and len(str(value)) > 50:
        value = str(value)[:50] + "..."
    print(f"{col_letter} ({header}): {value}")

# Check if column 11 has any data
print("\n" + "="*80)
print("COLUMN K (11) 'Refactorings Detected by tool' - First 10 rows:")
print("="*80)

for row_idx in range(2, 12):
    commit_hash = ws.cell(row=row_idx, column=2).value  # Column B
    col11_value = ws.cell(row=row_idx, column=11).value  # Column K
    if col11_value and len(str(col11_value)) > 70:
        col11_value = str(col11_value)[:70] + "..."
    print(f"Row {row_idx} - Commit: {str(commit_hash)[:12]} - Tool: {col11_value}")

# Count how many rows have data in column 11
count_with_data = 0
count_empty = 0
for row_idx in range(2, ws.max_row + 1):
    col11_value = ws.cell(row=row_idx, column=11).value
    if col11_value and str(col11_value).strip():
        count_with_data += 1
    else:
        count_empty += 1

print("\n" + "="*80)
print(f"Column K Summary: {count_with_data} rows with data, {count_empty} rows empty")
print("="*80)
